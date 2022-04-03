from random import randint
from math import floor
from typing import Union
from typing import Any
from pathlib import Path
from PIL import Image
import openpyxl


shade_number: int = 16
pix_amnt: int = 5
pix_len: int = 100


def number_to_list(decimal: int) -> list:
    """Form a based_list from decimal number according to a shade_number.

    Args:
        decimal (int): Decimal number will be converted to the based_list.
        
    List length: 
        The length is 3*pix_amnt*pix_amnt.

    Raises:
        TypeError: Decimal number must be a positive integer or a zero.

    Returns:
        list: based_list that consists numbers between 0 and shade_number-1.
    """
    if not isinstance(decimal, int) or decimal < 0:
        raise TypeError("Input data is not a positive integer or a zero")

    based_list = []
    
    for i in range(3*pix_amnt*pix_amnt):

        based_list.append(decimal % shade_number)
        decimal = decimal // shade_number

    # остатки от деления в обратном порядке по правилу перевода м/у системами:
    based_list.reverse()

    return based_list


def number_to_image(number_string: str) -> Image.Image:
    """Form an image from decimal number according to a shade_number.

    Args:
        number_string (str): A string that includes a number that will be
        converted to integer. Then integer will be converted to image.
        
    Library functions inside:
        number_to_list: To create a based_list from number that was converted
        from number_string.

    Raises:
        TypeError: Input data must be a string.

    Returns:
        Image.Image: A formed colored image from number with 
        size = (pix_amnt*pix_len, pix_amnt*pix_len).
    """
    if not isinstance(number_string, str):
        raise TypeError("Input data is not a string")

    number = int(number_string)
    based_list = number_to_list(number)
    size = (pix_amnt*pix_len, pix_amnt*pix_len)
    image = Image.new("RGB", size, (255, 255, 255))

    for x in range(pix_amnt*pix_len):
        for y in range(pix_amnt*pix_len):
            
            # big pixels position:
            x1 = floor(x/pix_len)
            y1 = floor(y/pix_len)

            # pixel_color = (c1, c2, c3):
            c1 = floor(255/shade_number*based_list[3*(pix_amnt*x1+y1)+0])
            c2 = floor(255/shade_number*based_list[3*(pix_amnt*x1+y1)+1])
            c3 = floor(255/shade_number*based_list[3*(pix_amnt*x1+y1)+2])

            image.putpixel((x, y), (c1, c2, c3))
    
    return image


def list_to_number(based_list: list) -> int:
    """Form a decimal number from based_list according to a shade_number.

    Args:
        based_list (list): List that consists numbers between 0 and shade_number-1. 
        Numbers depend on a shade_number.

    Raises:
        TypeError: Input data must be a list.

    Returns:
        int: A decimal number that was converted from notation with base = shadow_number
    """
    if not isinstance(based_list, list):
        raise TypeError("Input data is not a list")

    decimal: int = 0
    based_list.reverse()

    for index, value in enumerate(based_list):
        decimal += value * (shade_number**index)

    return decimal


def image_compress(input_image: Image.Image) -> Image.Image:
    """Compress the input image to a compressed_image.

    Args:
        input_image (Image.Image): An image that will be compressed.
        
    Input image size:
        The size is (pix_amnt*pix_len, pix_amnt*pix_len).

    Raises:
        TypeError: Input data must be a class PIL.Image.Image.

    Returns:
        Image.Image: A compressed_image with size = (pix_amnt, pix_amnt).
    """
    if not isinstance(input_image, Image.Image):
        raise TypeError("Input data is not a class PIL.Image.Image")

    size = (pix_amnt, pix_amnt)

    # create a blank image called "compressed":
    compressed_image = Image.new("RGB", size, (255, 255, 255))

    for x1 in range(pix_amnt):
        for y1 in range(pix_amnt):

            s1 = 0
            s2 = 0
            s3 = 0

            for x in range(pix_len):
                for y in range(pix_len):

                    (c1, c2, c3) = input_image.getpixel((x1*pix_len+x, y1*pix_len+y))
                    s1 += c1
                    s2 += c2
                    s3 += c3

            c1 = floor(s1/(pix_len*pix_len))
            c2 = floor(s2/(pix_len*pix_len))
            c3 = floor(s3/(pix_len*pix_len))

            compressed_image.putpixel((x1, y1), (c1, c2, c3))
    
    return compressed_image


def image_to_number(image: Image.Image) -> str:
    """Form a decimal number from an image according to a shade_number.

    Args:
        image (Image.Image): An image that will be converted to a decimal number.
        
    Library functions inside:
        image_compress: To compress an image to decipher it.
        list_to_number: To form a decimal number from decipherred based_list.

    Raises:
        TypeError: Input data must be a class PIL.Image.Image.
        ValueError: Input image size must be a (pix_amnt*pix_len, pix_amnt*pix_len).

    Returns:
        str: A number that converted to a string.
    """
    if not isinstance(image, Image.Image):
        raise TypeError("Input data is not a class PIL.Image.Image")
    
    size = (pix_amnt*pix_len, pix_amnt*pix_len)
    
    if size != image.size:
        raise ValueError("Input image size is not suitable")
    
    based_list = []
    image = image_compress(image)

    for x in range(pix_amnt):
        for y in range(pix_amnt):

            (c1, c2, c3) = image.getpixel((x, y))

            based_list.append(c1)
            based_list.append(c2)
            based_list.append(c3)
    
    for i in range(3*pix_amnt*pix_amnt):
        based_list[i] = round(based_list[i] * shade_number / 255)

    decimal = list_to_number(based_list)

    return str(decimal)


def search_row(book_path: str, identificator: str) -> int:
    """Finds a row number in the database (excel) and returns it.

    Args:
        book_path (str): Path to the excel book.
        identificator (str): The key that the row number must be found by.

    Returns:
        int: A row number in database (excel); 
        int: -1 otherwise.
    """
    book = openpyxl.load_workbook(book_path)
    sheet = book.active

    for i in range(1, 1000):
        if sheet.cell(row=i, column=1).value == identificator:
            return i

    return -1


def show_data(book_path: str, input_image: Image.Image) -> Union[str, int, None]:
    """Deciphers the image and returns the data from database (excel).
    If there is no data, returns None.

    Args:
        book_path (str): Path to the excel book.
        input_image (Image.Image): An image that will be decipherred.
        
    Library functions inside:
        image_to_number: To get an indentificator from image.
        search_row: To find the row in the database.

    Raises:
        TypeError: Input data must be a class PIL.Image.Image.

    Returns:
        str | int | None: The data that was found in database.
    """
    if not isinstance(input_image, Image.Image):
        raise TypeError("Input data is not a class PIL.Image.Image")
    
    book = openpyxl.load_workbook(book_path)
    sheet = book.active

    identificator = image_to_number(input_image)

    excel_row = search_row(book_path, identificator)

    if excel_row == -1: 
        return 

    return sheet.cell(row=excel_row, column=2).value


def id_generator(book_path: str) -> str:
    """Generates an identificator and returns it as a string.

    Args:
        book_path (str): Path to the excel book.

    Returns:
        str: An identificator as a string.
    """
    number = randint(0, shade_number**(3*pix_amnt*pix_amnt))

    while search_row(book_path, number) != -1:
        number = randint(0, shade_number**(3*pix_amnt*pix_amnt))
    
    return str(number)


def data_insert(book_path: str, data: Any) -> Image.Image:
    """Inserts the data into database (excel) and returns the created image.

    Args:
        book_path (str): Path to the excel book.
        data (Any): Data that will be inserted in database.
        
    Library functions inside:
        id_generator: To generate identificator.
        number_to_image: To create an image from number.

    Raises:
        RuntimeError: Error that can appear when database has not empty cells.

    Returns:
        Image.Image: An image with size = (pix_amnt*pix_len, pix_amnt*pix_len)
    """
    book = openpyxl.load_workbook(book_path)
    sheet = book.active

    i = 1
    generated_id = id_generator(book_path)

    while sheet.cell(row=i, column=1).value != " ":
        i += 1
        if i == 1005:
            raise RuntimeError("Can not find empty cell")
    
    sheet.cell(row=i, column=1).value = generated_id
    sheet.cell(row=i, column=2).value = data
    book.save(book_path)

    image = number_to_image(generated_id)
    
    return image


def message_file_open(message_file_name: str) -> str:
    message_file = open(message_file_name, "r")
    text_messages = message_file.read()
    message_file.close()
    return text_messages


def buttons_insert(book_path: str, data: list) -> None:
    """Inserts data to the book located at book_path.

    Args:
        book_path (str): Path to the excel book.
        data (list): Data to be inserted into excel book.

    Raises:
        RuntimeError: Error that can appear when excel book has not empty cells.
        
    Examples:
        :data = [identificator, yes_button, show_button]
        :data = [chat_id, False, False]
        :path_to_book = /path/buttons.xlsx
        :buttons_insert(path_to_book, data)
    """
    book = openpyxl.load_workbook(book_path)
    sheet = book.active
    
    excel_row = 1
    while sheet.cell(row=excel_row, column=1).value != " ":
        excel_row += 1
        if excel_row == 1005:
            raise RuntimeError("Can not find empty cell")
    
    data_len = len(data)
    excel_col = 1
    
    for d_index in range(data_len):
        sheet.cell(row=excel_row, column=excel_col).value = data[d_index]
        excel_col += 1
        
    book.save(book_path)
    
    return


def buttons_status(book_path: str, chat_id: int) -> Union[list, None]:
    """Returns the row from excel book by chat_id.

    Args:
        book_path (str): Path to the excel book.
        chat_id (int): telebot.message.chat.id

    Returns:
        Union[list, None]: The row from excel book or None if there is no any.
    """
    book = openpyxl.load_workbook(book_path)
    sheet = book.active
    
    excel_row = search_row(book_path, chat_id)
    
    if excel_row == -1:
        return
    
    data = []
    data.append(chat_id)
    data.append(sheet.cell(row=excel_row, column=2).value)
    data.append(sheet.cell(row=excel_row, column=3).value)
    
    return data


def buttons_change(book_path: str, data: list) -> Union[int, None]:
    """Changes the row in the excel book.

    Args:
        book_path (str): Path to the excel book.
        data (list): Data to be inserted instead of old in the excel book.

    Returns:
        Union[int, None]: None if data is changed. -2 otherwise.
    """
    book = openpyxl.load_workbook(book_path)
    sheet = book.active
    
    excel_row = search_row(book_path, data[0])
    
    if excel_row == -1:
        return -2
    
    sheet.cell(row=excel_row, column=2).value = data[1]
    sheet.cell(row=excel_row, column=3).value = data[2]
    book.save(book_path)
    
    return


def buttons_reset(book_path: str) -> None:
    """Clears 3 columns in the excel book.

    Args:
        book_path (str): Path to the excel book.
    """
    book = openpyxl.load_workbook(book_path)
    sheet = book.active
    
    for i in range(2, 1000):
        sheet.cell(row=i, column=1).value = " "
        sheet.cell(row=i, column=2).value = " "
        sheet.cell(row=i, column=3).value = " "

    book.save(book_path)

    return
